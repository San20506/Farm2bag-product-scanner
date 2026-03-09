"""
Excel reporter for generating detailed price comparison reports.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from typing import Dict, Any, List
from datetime import datetime
import os
from loguru import logger


class ExcelReporter:
    """
    Generates Excel reports with detailed product comparisons and summary statistics.
    """
    
    def __init__(self, config: Dict[str, Any] = None):
        """
        Initialize Excel reporter.
        
        Args:
            config: Configuration dictionary
        """
        self.config = config or {}
        self.report_dir = self.config.get('report_dir', 'data/reports')
        
        # Ensure report directory exists
        os.makedirs(self.report_dir, exist_ok=True)
    
    def generate_report(self, comparison_results: Dict[str, Any], 
                       output_filename: str = None) -> str:
        """
        Generate comprehensive Excel report from comparison results.
        
        Args:
            comparison_results: Results from PriceComparator
            output_filename: Custom filename (optional)
            
        Returns:
            Path to generated report file
        """
        if not output_filename:
            date_str = datetime.now().strftime('%Y-%m-%d')
            output_filename = f"{date_str}_report.xlsx"
        
        output_path = os.path.join(self.report_dir, output_filename)
        
        # Create workbook and worksheets
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        summary_sheet = workbook.create_sheet("Summary")
        detailed_sheet = workbook.create_sheet("Detailed Comparison")
        no_matches_sheet = workbook.create_sheet("No Matches Found")
        statistics_sheet = workbook.create_sheet("Statistics")
        
        # Generate each sheet
        self._create_summary_sheet(summary_sheet, comparison_results)
        self._create_detailed_sheet(detailed_sheet, comparison_results['matches'])
        self._create_no_matches_sheet(no_matches_sheet, comparison_results['no_matches'])
        self._create_statistics_sheet(statistics_sheet, comparison_results['statistics'], 
                                    comparison_results['price_analysis'])
        
        # Save workbook
        workbook.save(output_path)
        logger.info(f"Excel report saved to: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, sheet, results: Dict[str, Any]):
        """Create executive summary sheet."""
        sheet.title = "Executive Summary"
        
        # Title
        sheet['A1'] = "Product Price Comparison Report"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Key metrics
        stats = results.get('statistics', {})
        
        sheet['A4'] = "Key Metrics"
        sheet['A4'].font = Font(size=14, bold=True)
        
        metrics = [
            ("Total Products Compared", stats.get('total_matches', 0)),
            ("Source Cheaper", stats.get('source_cheaper_count', 0)),
            ("Target Cheaper", stats.get('target_cheaper_count', 0)),
            ("Source Cheaper (%)", f"{stats.get('source_cheaper_percentage', 0):.1f}%"),
            ("Average Price Difference", f"{stats.get('average_price_difference_percentage', 0):.1f}%"),
            ("Maximum Savings", f"{abs(stats.get('max_savings_percentage', 0)):.1f}%"),
        ]
        
        row = 5
        for metric, value in metrics:
            sheet[f'A{row}'] = metric
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Top savings opportunities
        matches = results.get('matches', [])
        if matches:
            sheet['A12'] = "Top 5 Savings Opportunities"
            sheet['A12'].font = Font(size=14, bold=True)
            
            # Sort by biggest savings (most negative percentage difference)
            top_savings = sorted(matches, key=lambda x: x['price_comparison']['percentage_difference'])[:5]
            
            headers = ['Product', 'Source Site', 'Source Price', 'Target Site', 'Target Price', 'Savings %']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=13, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, match in enumerate(top_savings, 14):
                sheet.cell(row, 1, match['source_product']['name'])
                sheet.cell(row, 2, match.get('source_site', 'Unknown'))
                sheet.cell(row, 3, f"₹{match['price_comparison']['source_price']:.2f}")
                sheet.cell(row, 4, match.get('target_site', 'Unknown'))
                sheet.cell(row, 5, f"₹{match['price_comparison']['target_price']:.2f}")
                sheet.cell(row, 6, f"{abs(match['price_comparison']['percentage_difference']):.1f}%")
        
        # Auto-adjust column widths
        for col in range(1, 7):
            sheet.column_dimensions[chr(64 + col)].width = 20
    
    def _create_detailed_sheet(self, sheet, matches: List[Dict[str, Any]]):
        """Create detailed comparison sheet."""
        if not matches:
            sheet['A1'] = "No matches found"
            return
        
        headers = [
            'Source Product', 'Source Site', 'Source Price', 'Source Unit Price',
            'Target Product', 'Target Site', 'Target Price', 'Target Unit Price',
            'Price Difference (₹)', 'Price Difference (%)', 'Unit Price Diff (%)',
            'Price Advantage', 'Similarity Score', 'Category'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add data
        for row, match in enumerate(matches, 2):
            source = match['source_product']
            target = match['target_product']
            price_comp = match['price_comparison']
            unit_comp = match['unit_price_comparison']
            
            data = [
                source['name'],
                match.get('source_site', 'Unknown'),
                price_comp['source_price'],
                unit_comp['source_price_per_unit'],
                target['name'],
                match.get('target_site', 'Unknown'),
                price_comp['target_price'],
                unit_comp['target_price_per_unit'],
                price_comp['absolute_difference'],
                price_comp['percentage_difference'],
                unit_comp['per_unit_percentage'],
                price_comp['price_advantage'],
                match['similarity_score'],
                source.get('normalized_category', 'Unknown')
            ]
            
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Color coding for price advantage
                if col in [10, 11]:  # Price difference columns
                    if isinstance(value, (int, float)) and value < 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    elif isinstance(value, (int, float)) and value > 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 15
    
    def _create_no_matches_sheet(self, sheet, no_matches: List[Dict[str, Any]]):
        """Create sheet for products with no matches."""
        if not no_matches:
            sheet['A1'] = "All products have matches!"
            return
        
        headers = ['Product Name', 'Site', 'Category', 'Brand', 'Price', 'Reason']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add data
        for row, item in enumerate(no_matches, 2):
            product = item['product']
            data = [
                product.get('name', 'Unknown'),
                item.get('site', product.get('site', 'Unknown')),
                product.get('normalized_category', 'Unknown'),
                product.get('normalized_brand', 'Unknown'),
                product.get('price', 0),
                item.get('reason', 'No reason specified')
            ]
            
            for col, value in enumerate(data, 1):
                sheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[chr(64 + col)].width = 20
    
    def _create_statistics_sheet(self, sheet, statistics: Dict[str, Any], 
                               price_analysis: Dict[str, Any]):
        """Create statistics and analysis sheet."""
        sheet['A1'] = "Detailed Statistics & Analysis"
        sheet['A1'].font = Font(size=16, bold=True)
        
        # Overall statistics
        sheet['A3'] = "Overall Statistics"
        sheet['A3'].font = Font(size=14, bold=True)
        
        stats_data = [
            ("Total Matches", statistics.get('total_matches', 0)),
            ("Source Cheaper Count", statistics.get('source_cheaper_count', 0)),
            ("Target Cheaper Count", statistics.get('target_cheaper_count', 0)),
            ("Source Competitive Rate", f"{statistics.get('source_cheaper_percentage', 0):.1f}%"),
            ("Average Price Difference", f"{statistics.get('average_price_difference_percentage', 0):.1f}%"),
            ("Median Price Difference", f"{statistics.get('median_price_difference', 0):.1f}%"),
            ("Maximum Savings Opportunity", f"{abs(statistics.get('max_savings_percentage', 0)):.1f}%"),
            ("Minimum Savings", f"{statistics.get('min_savings_percentage', 0):.1f}%")
        ]
        
        row = 4
        for stat, value in stats_data:
            sheet[f'A{row}'] = stat
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Category analysis
        if 'by_category' in price_analysis:
            sheet['A13'] = "Performance by Category"
            sheet['A13'].font = Font(size=14, bold=True)
            
            cat_headers = ['Category', 'Avg Price Diff %', 'Product Count', 'Source Cheaper Count']
            for col, header in enumerate(cat_headers, 1):
                cell = sheet.cell(row=14, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Site analysis
        if 'by_site' in price_analysis:
            sheet['A20'] = "Performance by Site"
            sheet['A20'].font = Font(size=14, bold=True)
            
            site_headers = ['Site', 'Avg Price Diff %', 'Product Count', 'Source Cheaper Count']
            for col, header in enumerate(site_headers, 1):
                cell = sheet.cell(row=21, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for col in range(1, 5):
            sheet.column_dimensions[chr(64 + col)].width = 25